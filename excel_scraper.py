from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference
from scraper_settings import ScraperSettings

def get_data (workbook: Workbook, data_sheet_name: str) -> list:
    """
    Function to scrape data from configured excel worksheet

    :param workbook         : Workbook object created based on scraper settings
    :param data_sheet_name  : Name of worksheet that contains data to be scraped
    """
    data_sheet = workbook[data_sheet_name]

    items_col1 = []
    items_col2 = []
    for row in data_sheet:
        # Get value of first cell in each row
        items_col1.append(row[0].value)
        # Get value of second cell in each row
        items_col2.append(row[1].value)

    return construct_data(col1=items_col1, col2=items_col2)

def construct_data (col1: list, col2: list) -> list:
    """
    Constructs data from excel worksheet into 2D array

    :param col1     : First column of scraped data
    :param col2     : Second column of scraped data
    """
    data = []
    for i in range(len(col1)):
        # Create an entry for each row in the data sheet
        data.append([col1[i], col2[i]])
    
    print(data)
    return data

def generate_line_graph (workbook: Workbook, workbook_name: str, graph_sheet_name: str, data: list):
    """
    Constructs line graphs from data pulled from an excel worksheet
    
    :param workbook         : Workbook object created based on scraper settings
    :param workbook_name    : Name of workbook 
    :param graph_sheet_name : Name of sheet that graphs will be placed
    :param data             : Data scraped from different excel sheet
    """
    # Clear worksheet to overwrite old data/graph
    del workbook[graph_sheet_name]
    workbook.create_sheet(title=graph_sheet_name)
    worksheet = workbook[graph_sheet_name]

    for row in data:
        worksheet.append(row)

    chart = LineChart()
    chart.title = 'line chart'
    chart.style = 13
    chart.y_axis.title = 'Size'
    chart.x_axis.title = 'Test Number'

    data = Reference(worksheet=worksheet, min_col=2, min_row=1, max_col=4, max_row=7)
    chart.add_data(data=data, titles_from_data=True)

    worksheet.add_chart(chart, "A10")

    workbook.save(workbook_name)

def main ():
    """
    Main function that runs the scraper logic to scrape data and generate line graph
    """
    settings = ScraperSettings()
    # Extract contens from settings
    file = settings.get_path_to_workbook()
    workbook = load_workbook(filename=file)
    workbook_name = settings.XLSX_FILE_NAME
    data_sheet = settings.DATA_SHEET
    graph_sheet = settings.GRAPH_SHEET

    constructed_data = get_data(workbook=workbook, data_sheet_name=data_sheet)

    generate_line_graph(workbook=workbook, workbook_name=workbook_name, graph_sheet_name=graph_sheet, data=constructed_data)

main()